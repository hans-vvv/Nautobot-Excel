import openpyxl
from pynautobot import api


url = "http://192.168.248.144:8080"
token = "31988337481f153797bf052d28291987042d717b"

nautobot = api(url=url, token=token)

wb = openpyxl.load_workbook("Nautobot-upload.xlsx")

fk_name_exceptions = {
    "dcim_device_types": "model",
    "ipam_ip-address": "address",
    "ipam_prefixes": "prefix",
    "ipam_aggregates": "prefix",
    "ipam_vlans": "vid",
    "circuits_circuits": "cid",
}  

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]

    columnnames = [column[0].value.strip() for column in sheet.iter_cols(0, sheet.max_column)]
    
    objects_properties = []
    for index, row in enumerate(sheet.iter_rows()):
        if index == 0:
            continue
        table_row = [str(cell.value).strip() if str(cell.value).strip() != "None" else None 
                     for cell in row]

        object_properties = {}
        for columnname, cell in zip(columnnames, table_row):
            if cell is None:
                continue
            if ":" in columnname: # Foreign key
                local_columnname, foreign_tablename = columnname.split(":")[0], columnname.split(":")[1]
                if foreign_tablename in fk_name_exceptions.keys():
                    value = {fk_name_exceptions[foreign_tablename]: cell}
                else:
                    value = {"name": cell}
                object_properties.update({local_columnname: value})
            else:
                object_properties.update({columnname: cell})
        objects_properties.append(object_properties)
    # Dump objects in dB
    first_field, second_field = sheetname.split(".")[0], sheetname.split(".")[1]
    objects_type = getattr(getattr(nautobot, first_field), second_field)
    objects_type.create(objects_properties)

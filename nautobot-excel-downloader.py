import json
import openpyxl
import requests
from pynautobot import api
from collections import defaultdict
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter


def xlref(row, column, zero_indexed=True):
    """
    openpyxl helper
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


url = "http://192.168.248.144:8080"
token = "31988337481f153797bf052d28291987042d717b"

nautobot = api(url=url, token=token)

# Discover relevant API endpoints using requests module
response = requests.get(url + '/api')
nautobot_api_primary_fields = [key for key in response.json().keys()
                              if (key != "graphql" and key != "status" and key != "extras")]
nautobot_api_endpoints = {}
for primary_field in nautobot_api_primary_fields:
    response = requests.get(url + '/api/' + primary_field)
    nautobot_api_secondary_fields = list(response.json().keys())
    nautobot_api_endpoints.update({primary_field: nautobot_api_secondary_fields})

# Dump tables in memory
nautobot_data = defaultdict(lambda: defaultdict(list))
for primary_field in nautobot_api_endpoints:
    for secondary_field in nautobot_api_endpoints[primary_field]:
        objects_type = getattr(getattr(nautobot, primary_field), secondary_field)
        try:
            for object_ in objects_type.all():
                nautobot_data[primary_field][secondary_field].append(object_.serialize())
        except:
            pass

# print(json.dumps(nautobot_data, indent=4))

# Print to Excel
wb = Workbook()
for primary_field in nautobot_data:
    for secondary_field in nautobot_data[primary_field]:
        wb.create_sheet(primary_field + "." + secondary_field)
        sheet = wb[primary_field + "." + secondary_field]
        # print Column headers
        for index, key in enumerate(nautobot_data[primary_field][secondary_field][0].keys()):
            sheet[xlref(0, index)] = key
        # print Data
        row_index = 1
        for object_ in nautobot_data[primary_field][secondary_field]:
            for col_index, val in enumerate(object_.values()):
                if not val:
                    sheet[xlref(row_index, col_index)] = ""
                elif isinstance(val, list):
                    val = [str(element) for element in val]
                    sheet[xlref(row_index, col_index)] = ','.join(val)
                else:
                    sheet[xlref(row_index, col_index)] = str(val)
            row_index += 1

wb.save('Nautobot-tables.xlsx')
